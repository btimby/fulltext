from io import StringIO
import datetime

import openpyxl

from fulltext.util import BaseBackend
from fulltext.util import assert_cmd_exists, exiftool_title


class Backend(BaseBackend):

    def check(self, title):
        if title:
            assert_cmd_exists('exiftool')

    def handle_fobj(self, fobj):
        text = StringIO()
        wb = openpyxl.load_workbook(fobj)
        for n in wb.sheetnames:
            ws = wb[n]
            for x in range(1, ws.max_row+1):
                for y in range(1, ws.max_column+1):
                    v = ws.cell(x, y).value
                    if v:
                        if isinstance(v, (int, float, datetime.datetime)):
                            v = str(v)
                        text.write(v)
                        text.write(u'\t')
                text.write(u'\n')
        return text.getvalue()
    
    def handle_path(self, path):
        with open(path, 'rb') as fin:
            return self.handle_fobj(fin)

    def handle_title(self, f):
        return exiftool_title(f, self.encoding, self.encoding_errors)
    
